"""Validation logic for the Namescan emerald API."""
import csv
import dataclasses
import glob
import json
import sys
from dataclasses import dataclass
from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Optional, Any, Tuple, TextIO

import openpyxl
import requests
from click import BadParameter
from openpyxl.worksheet.worksheet import Worksheet
from requests import Response
from rich import prompt
from rich.console import Console
from rich.markdown import Markdown

from models import (
    EntityToScan,
    OrganizationToScan,
    PersonToScan,
    Entity,
    PersonScanResult,
    OrganisationScanResult,
)

NAME_SCAN_URL = "https://api.namescan.io/v3"
EMERALD_PERSON_URL = f"{NAME_SCAN_URL}/person-scans/emerald"
EMERALD_ORGANIZATION_URL = f"{NAME_SCAN_URL}/organisation-scans/emerald"
REQUEST_TIMEOUT_IN_SECONDS = 10


def log_request(request_body: dict, output_file: Path):
    """Log a request to a file."""
    with open(output_file, "w", encoding="utf-8") as file:
        file.write(json.dumps(request_body, indent=4))


def response_age(response: dict) -> int:
    """Get the age of a response in days."""
    date = response["date"]
    response_date = datetime.fromisoformat(date)
    return (datetime.now().astimezone() - response_date).days


def file_response(file_path: Path, max_days_old: int) -> Optional[Tuple[int, Response]]:
    if not file_path.exists():
        return None

    response = Response()
    response.status_code = 201
    response._content = bytes(  # pylint: disable=protected-access
        file_path.read_text("utf-8"), "utf-8"
    )
    age = response_age(response.json())
    if age > max_days_old:
        return None

    return age, response


def send_request(
    console: Console,
    entity: EntityToScan,
    index: str,
    api_url: str,
    entity_dict: dict,
    key: str,
    output_path: Path,
    max_days_old: int,
) -> None:
    """Send a request to the Namescan emerald API."""
    status_prefix = f"{index} checking {entity.name}..."
    with console.status(status_prefix) as status:
        log_request(entity_dict, Path(output_path, f"{entity.hash}.req.json"))
        output_file = Path(output_path, f"{entity.hash}.resp.json")
        age, response = get_response(
            api_url, entity_dict, key, max_days_old, output_file
        )

        if response.status_code < 300:
            response_json = response.json()
            when = "checked just now" if age == 0 else f"checked {age} days ago"
            status.console.log(
                f"{index} {when} {entity.name} - {response_json.get('numberOfMatches', 'Error')} matches"
            )
            log_request(response_json, output_file)
        else:
            console.log(
                f"[red]Error while sending request {index} {entity.hash}, {entity.name} "
                f"to Namescan API: {response.status_code}"
                f" - {response.text}[/red]"
            )
            if not prompt.Confirm.ask(f"Continue after row {index} ?"):
                console.log("Aborted")
                sys.exit(0)


def get_response(
    api_url, entity_dict, key, max_days_old, output_file
) -> Tuple[int, Response]:
    from_file = file_response(output_file, max_days_old)
    if from_file is not None:
        return from_file

    return (
        0,
        requests.post(
            api_url,
            json=entity_dict,
            headers={"api-key": key},
            timeout=REQUEST_TIMEOUT_IN_SECONDS,
        ),
    )


def convert_to_csv(worksheet: Worksheet):
    """Convert an Excel file to a CSV file."""
    csv_file = StringIO()
    writer = csv.writer(
        csv_file, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL
    )
    for row in worksheet.iter_rows(values_only=True):
        writer.writerow(row)
    csv_file.seek(0)
    return csv_file


def read_as_dataframe(file: Path) -> list[dict[str, Any]]:
    extension = file.suffix
    list_of_dicts = []
    worksheet: Worksheet = (
        read_csv_file_as_worksheet(file)
        if extension == ".csv"
        else read_csv_as_worksheet(
            convert_to_csv(openpyxl.load_workbook(file).worksheets[0])
        )
    )
    headers = [str(header.value) for header in worksheet[1] if header.value]

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        res = dict(zip(headers, list(row)))
        list_of_dicts.append(res)
    return list_of_dicts


def read_csv_file_as_worksheet(file_path: Path) -> Worksheet:
    with open(file_path, encoding="utf-8") as csv_file:
        return read_csv_as_worksheet(csv_file)


def read_csv_as_worksheet(csv_file: TextIO) -> Worksheet:
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]
    reader = csv.reader(csv_file)
    for row in reader:
        worksheet.append(row)
    return worksheet


def check_database(
    console: Console, path: Path, max_age: int, entities: list[EntityToScan]
) -> None:
    """Check the contents of the cache."""

    files = glob.glob(f"{path}/*.resp.json")
    response_dict = {
        Path(file).name.split(".")[0]: json.loads(Path(file).read_text("utf-8"))
        for file in files
    }
    total_outdated = 0
    in_dictionary = 0
    not_in_dictionary = 0

    for entity in entities:
        if entity.hash in response_dict:
            in_dictionary += 1
            entity_dict = response_dict[entity.hash]
            age = response_age(entity_dict)
            if age > max_age:
                total_outdated += 1
        else:
            not_in_dictionary += 1

    total = len(entities)
    cache_size = len(response_dict)
    question = (
        f"[bold]{total}[/bold] rows in input. Found [bold]{len(response_dict)}[/bold] responses in"
        f" [italic]{path}[/italic], [bold]{in_dictionary}[/bold] of them match rows in the input file. \n"
        f"[bold]{total_outdated}[/bold] of the responses in the cache are more than {max_age} days old. \n"
        f"Will call namescan [bold red]{not_in_dictionary + total_outdated}[/bold red] times. Continue?"
        if cache_size > 0
        else f"Found no responses in [italic]{path}[/italic].\nWill call namescan "
        f"[bold magenta]{total}[/bold magenta] times. Once for each row in the excel sheet. Continue?"
    )

    if not prompt.Confirm.ask(question):
        console.log("Aborted")
        sys.exit(0)


def to_entities(entity: str, dataframe: list[dict[str, Any]]) -> list[EntityToScan]:
    return [
        OrganizationToScan.from_dataframe(row)
        if entity == "organization"
        else PersonToScan.from_dataframe(row)
        for _, row in enumerate(dataframe)
    ]


def validate_file(
    console: Console,
    entities: list[EntityToScan],
    output_path: Path,
    key: str,
    max_days_old: int,
) -> None:
    """Validate an Excel sheet with persons against the Namescan emerald API."""

    output_path.mkdir(parents=True, exist_ok=True)

    for index, entity in enumerate(entities):
        if isinstance(entity, OrganizationToScan):
            send_request(
                console,
                entity,
                str(index),
                EMERALD_ORGANIZATION_URL,
                dataclasses.asdict(entity),
                key,
                output_path,
                max_days_old,
            )
        elif isinstance(entity, PersonToScan):
            send_request(
                console,
                entity,
                str(index),
                EMERALD_PERSON_URL,
                dataclasses.asdict(entity),
                key,
                output_path,
                max_days_old,
            )
        else:
            raise BadParameter(f"Unknown scan type: {entity}")


@dataclass(frozen=True)
class Rationale:
    last_updated: datetime
    entity_to_scan: EntityToScan
    matches_with_explanations: dict[Entity, Optional[str]]

    @property
    def matches(self) -> int:
        return len(self.matches_with_explanations.keys())

    @property
    def explained(self) -> int:
        return len(
            [
                explanation
                for explanation in self.matches_with_explanations.values()
                if explanation is not None
            ]
        )

    @property
    def rationale(self):
        return ", ".join(
            [
                f"{entity.name}: {explanation}"
                for entity, explanation in self.matches_with_explanations.items()
                if explanation is not None
            ]
        )

    @property
    def no_rationale(self):
        return ", ".join(
            [
                f"{entity.name}: {entity.entity_summary}"
                for entity, explanation in self.matches_with_explanations.items()
                if explanation is None
            ]
        )

    @property
    def icon(self):
        return "🟢" if self.matches == self.explained else "🔴"


def to_matrix(
    dataframe, last_updated, explanation, matched, need_explanation, unique_id, verdict
) -> list[list[str]]:
    results = []
    for index, row in enumerate(dataframe):
        with_it = list(row.values()) + [
            unique_id[index],
            last_updated[index],
            matched[index],
            verdict[index],
            explanation[index],
            need_explanation[index],
        ]
        results.append(list(with_it))
    return results


def add_rationale(  # pylint: disable=too-many-locals
    console: Console,
    input_file: Path,
    entity: str,
    database_path: Path,
    output_sheet: Path,
    file_format: str = ".xlsx",
) -> None:
    console.log(Markdown(f"Writing `{output_sheet}`"))
    dataframe = read_as_dataframe(input_file)

    entities = [
        (
            index,
            PersonToScan.from_dataframe(row)
            if entity == "person"
            else OrganizationToScan.from_dataframe(row),
        )
        for index, row in enumerate(dataframe)
    ]

    rationales: list[Rationale] = [
        create_rationale(
            console,
            str(index),
            entity,
            entity.hash,
            database_path,
        )
        for index, entity in entities
    ]

    def to_verdict(rationale: Rationale) -> str:
        if rationale.explained == rationale.matches:
            return "False positive"
        if rationale.matches == 0:
            return "No match"
        return "Needs explanation"

    unique_id = [rationale.entity_to_scan.hash for rationale in rationales]
    last_checked = [
        rationale.last_updated.strftime("%m/%d/%Y") for rationale in rationales
    ]
    matched = [rationale.matches > 0 for rationale in rationales]
    verdict = [to_verdict(rationale) for rationale in rationales]
    explanation = [rationale.rationale for rationale in rationales]
    need_explanation = [rationale.no_rationale for rationale in rationales]

    headers = list(dataframe[0].keys()) + [
        "UniqueId",
        "LastChecked",
        "Matched",
        "Verdict",
        "Explanation",
        "NeedExplanation",
    ]

    matrix = to_matrix(
        dataframe,
        last_checked,
        explanation,
        matched,
        need_explanation,
        unique_id,
        verdict,
    )

    if file_format == ".xlsx":
        write_excel_sheet(
            output_sheet,
            headers,
            matrix,
        )
    else:
        write_csv(output_sheet, headers, matrix)

    total_matches = sum(rationale.matches for rationale in rationales)
    total_explained = sum(rationale.explained for rationale in rationales)
    console.log(f"Total matches: {total_matches}, total explained: {total_explained}")


def write_excel_sheet(
    output_sheet,
    headers,
    matrix,
) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.worksheets[0]
    worksheet.append(headers)
    for row in matrix:
        worksheet.append(row)
    workbook.save(output_sheet)


def write_csv(
    output_sheet,
    headers,
    matrix,
):
    with open(output_sheet, "w", newline="", encoding="utf-8") as csvfile:
        writer = csv.writer(
            csvfile, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL
        )
        writer.writerow(headers)
        for row in matrix:
            writer.writerow(row)


def create_rationale(
    console: Console,
    index: str,
    entity: EntityToScan,
    person_hash: str,
    output_path: Path,
) -> Rationale:
    response_json_string = Path(output_path, f"{person_hash}.resp.json").read_text(
        encoding="utf-8"
    )
    json_object = json.loads(response_json_string)
    try:
        scan_result = (
            PersonScanResult.from_json(json_object)
            if isinstance(entity, PersonToScan)
            else OrganisationScanResult.from_json(json_object)
        )
    except (KeyError, ValueError) as exc:
        console.log(
            f"⚠️ Error while parsing response {index} {person_hash}.resp.json, {entity.name} "
            f"from Namescan API: {exc}"
        )
        return Rationale(datetime.now(), entity, {})

    rationale = Rationale(
        last_updated=datetime.fromisoformat(scan_result.date),
        entity_to_scan=entity,
        matches_with_explanations={
            match: match.rationale for match in scan_result.entities
        },
    )
    console.log(
        f"{rationale.icon} {index} {entity.name} -> {rationale.matches} matches,"
        f" {rationale.explained} false positive."
    )
    return rationale
